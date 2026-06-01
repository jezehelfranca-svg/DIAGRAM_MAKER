import os
import sys
import json
import zlib
import base64
from flask import Flask, request, jsonify, render_template, send_file
import requests

project_dir = os.path.dirname(os.path.abspath(__file__))
app = Flask(__name__, template_folder=os.path.join(project_dir, 'templates'))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB limit

# ----------------- Navigation Routes -----------------

@app.route('/')
def diagram_maker():
    """Serves the Diagram Maker dashboard as the primary landing page."""
    return render_template('diagram_maker.html')

@app.route('/cable-routing')
def cable_routing():
    """Serves the Cable Routing Decision Navigator."""
    return render_template('cable_routing.html')

@app.route('/extractor')
def extractor():
    """Serves the Extractor UI."""
    return render_template('index.html')

# ----------------- Diagram Maker API Endpoints -----------------

@app.route('/api/diagrams', methods=['GET'])
def list_diagrams():
    """Scans the repository directory for local `.mmd` diagrams."""
    try:
        diagrams = []
        # Scan root folder for any .mmd files
        for f in os.listdir(project_dir):
            if f.lower().endswith('.mmd'):
                diagrams.append({
                    "name": f,
                    "path": f,
                    "location": "root"
                })
        return jsonify({"success": True, "diagrams": diagrams})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/diagrams/load', methods=['GET'])
def load_diagram():
    """Loads the raw content of a specific `.mmd` file."""
    rel_path = request.args.get('path', '').strip()
    if not rel_path or '..' in rel_path:
        return jsonify({"success": False, "error": "Invalid diagram path."}), 400
    
    full_path = os.path.join(project_dir, rel_path)
    if not os.path.exists(full_path) or not full_path.lower().endswith('.mmd'):
        return jsonify({"success": False, "error": "File not found."}), 404
        
    try:
        with open(full_path, "r", encoding="utf-8") as f:
            code = f.read()
        return jsonify({"success": True, "code": code})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/diagrams/save', methods=['POST'])
def save_diagram():
    """Saves modified Mermaid code back to the local `.mmd` file."""
    # Handle both multipart form data and raw JSON requests
    if request.is_json:
        req_data = request.get_json()
        rel_path = req_data.get('path', '').strip()
        code = req_data.get('code', '')
    else:
        rel_path = request.form.get('path', '').strip()
        code = request.form.get('code', '')
    
    if not rel_path or '..' in rel_path:
        return jsonify({"success": False, "error": "Invalid diagram path."}), 400
        
    full_path = os.path.join(project_dir, rel_path)
    if not full_path.lower().endswith('.mmd'):
        return jsonify({"success": False, "error": "File must have .mmd extension."}), 400
        
    try:
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        with open(full_path, "w", encoding="utf-8") as f:
            f.write(code)
        print(f"Diagram saved successfully to {full_path}")
        return jsonify({"success": True, "message": "Diagram saved successfully."})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/diagrams/export-pdf', methods=['POST'])
def export_diagram_pdf():
    """Compresses Mermaid syntax and retrieves a compiled A3 Landscape PDF via mermaid.ink."""
    # Handle both multipart form data and raw JSON requests
    if request.is_json:
        req_data = request.get_json()
        rel_path = req_data.get('path', '').strip()
        code = req_data.get('code', '')
    else:
        rel_path = request.form.get('path', '').strip()
        code = request.form.get('code', '')
    
    if not rel_path or '..' in rel_path:
        return jsonify({"success": False, "error": "Invalid diagram path."}), 400
        
    full_path = os.path.join(project_dir, rel_path)
    if not full_path.lower().endswith('.mmd'):
        return jsonify({"success": False, "error": "Invalid file format."}), 400
        
    pdf_path = full_path.replace('.mmd', '.pdf')
    
    try:
        # Strip comments and redundant blank lines to optimize URL length
        cleaned_lines = []
        for line in code.splitlines():
            stripped = line.strip()
            if stripped.startswith("%%") and not stripped.startswith("%%{"):
                continue
            if not stripped:
                continue
            cleaned_lines.append(stripped)
        cleaned_code = "\n".join(cleaned_lines)
        
        # Prepare state JSON for mermaid.ink
        state = {
            "code": cleaned_code,
            "mermaid": {"theme": "default"}
        }
        json_str = json.dumps(state, separators=(',', ':'))
        
        # Compress and base64 urlsafe encode the payload
        compressed = zlib.compress(json_str.encode('utf-8'), level=9)
        encoded = base64.b64encode(compressed).decode('utf-8')
        encoded_urlsafe = encoded.replace('+', '-').replace('/', '_').replace('=', '')
        
        url = f"https://mermaid.ink/pdf/pako:{encoded_urlsafe}?fit&landscape&paper=a3"
        print(f"Requesting A3 PDF export from: {url[:100]}...")
        
        headers = {"User-Agent": "Mozilla/5.0"}
        r = requests.get(url, headers=headers, timeout=60)
        
        if r.status_code == 200:
            with open(pdf_path, "wb") as f:
                f.write(r.content)
            print(f"Exported PDF saved to {pdf_path}")
            return send_file(pdf_path, mimetype='application/pdf', as_attachment=True, download_name=os.path.basename(pdf_path))
        else:
            return jsonify({
                "success": False, 
                "error": f"Failed to generate PDF from mermaid.ink. Status code: {r.status_code}",
                "detail": r.text[:200]
            }), 500
            
    except Exception as e:
        print(f"Error exporting PDF: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/cable-routing/export-excel', methods=['POST'])
def export_cable_routing_excel():
    """Populates inputs/outputs into Cable_Routing_Tool_Full_Logic.xlsx template and streams it."""
    from io import BytesIO
    from openpyxl import load_workbook
    
    try:
        data = request.get_json()
        inputs = data.get('inputs', {})
        outputs = data.get('outputs', {})
        
        template_path = os.path.join(project_dir, "Cable_Routing_Tool_Full_Logic.xlsx")
        if not os.path.exists(template_path):
            return jsonify({"success": False, "error": "Excel template not found."}), 404
            
        wb = load_workbook(template_path)
        ws = wb['Routing Decision Tool']
        
        # Populate inputs (Col B)
        ws['B4'] = inputs.get('area', '')
        ws['B5'] = inputs.get('primary', '')
        ws['B6'] = inputs.get('shared', '')
        ws['B7'] = inputs.get('same', '')
        ws['B8'] = inputs.get('sep', '')
        ws['B9'] = inputs.get('sched', '')
        try:
            ws['B10'] = float(inputs.get('spare', 20)) / 100.0
        except ValueError:
            ws['B10'] = 0.20
            
        # Output cells (B13-B20) are left unmodified to preserve the template's Excel formulas.
        # Excel will dynamically calculate these cells based on the inputs when opened.
        
        # Populate notes in A23 (replaces placeholder text)
        if inputs.get('notes'):
            ws['A23'] = inputs.get('notes')
            
        # Stream file back to client
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Cable_Routing_Decision_Output.xlsx"
        )
    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ----------------- Extractor Stub APIs -----------------
# Implemented to prevent index.html console errors when running standalone.

@app.route('/api/list-pdfs', methods=['GET'])
def list_pdfs():
    """Lists any local PDF files in the directory for the UI explorer shell."""
    try:
        files = sorted([f for f in os.listdir(project_dir) if f.lower().endswith('.pdf')])
        pdf_list = []
        for f in files:
            size_mb = os.path.getsize(os.path.join(project_dir, f)) / (1024 * 1024)
            pdf_list.append({
                "name": f,
                "size_mb": round(size_mb, 2),
                "has_mapping": False
            })
        return jsonify({"success": True, "pdfs": pdf_list})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

@app.route('/api/logs', methods=['GET'])
def get_logs():
    return jsonify({"success": True, "logs": ["System idle. Standalone UI Mode active."]})

@app.route('/api/status', methods=['GET'])
def get_status():
    return jsonify({
        "status": "idle",
        "count": 0,
        "error": None,
        "output_xlsx": "telecom_extracted_requirements.xlsx",
        "output_csv": "telecom_extracted_requirements.csv"
    })

if __name__ == '__main__':
    print("Starting Standalone Diagram Maker Web Application at http://127.0.0.1:5000...")
    app.run(host='127.0.0.1', port=5000, debug=False)
